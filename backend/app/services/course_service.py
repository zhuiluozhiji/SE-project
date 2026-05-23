import csv
import io
import re
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.course_schedule import CourseSchedule
from app.models.schedule_event import ScheduleEvent
from app.schemas.course import CourseCreate

DEFAULT_USER_ID = 1
COURSE_TITLE_SUFFIX = "课程"

SECTION_TIMES: dict[int, tuple[time, time]] = {
    1: (time(8, 0), time(8, 45)),
    2: (time(8, 50), time(9, 35)),
    3: (time(10, 0), time(10, 45)),
    4: (time(10, 50), time(11, 35)),
    5: (time(11, 40), time(12, 25)),
    6: (time(13, 25), time(14, 10)),
    7: (time(14, 15), time(15, 0)),
    8: (time(15, 5), time(15, 50)),
    9: (time(16, 15), time(17, 0)),
    10: (time(17, 5), time(17, 50)),
    11: (time(18, 50), time(19, 35)),
    12: (time(19, 40), time(20, 25)),
    13: (time(20, 30), time(21, 15)),
}

HEADER_MAP = {
    "course_code": {"course_code", "课程代码"},
    "course_name": {"course_name", "课程名", "课程名称", "name"},
    "teacher": {"teacher", "教师", "教师姓名", "授课教师", "老师"},
    "semester": {"semester", "学期"},
    "schedule_time": {"schedule_time", "上课时间", "上课时间地点", "时间"},
    "weekday": {"weekday", "星期", "周几", "day"},
    "start_section": {"start_section", "开始节次", "起始节次"},
    "end_section": {"end_section", "结束节次", "终止节次"},
    "sections": {"sections", "节次", "上课节次"},
    "weeks": {"weeks", "周次", "教学周"},
    "location": {"location", "地点", "教室", "上课地点"},
}

REQUIRED_COLUMNS = {"course_name", "weekday"}
SECTION_COLUMN_GROUPS = ({"sections"}, {"start_section", "end_section"})
ZJU_EXPORT_REQUIRED_COLUMNS = {"course_name", "schedule_time"}
COURSE_TEMPLATE_COLUMNS = ["课程名", "星期", "节次", "地点", "教师", "周次"]
COURSE_TEMPLATE_ROWS = [
    ["软件工程", "周二", "3-4", "玉泉曹楼", "李老师", "1-16"],
    ["数据库系统", "周三", "6-7", "紫金港东2", "王老师", "1-16"],
]
ZJU_TEMPLATE_TITLE = "2025-2026学年春夏学期某某的课表"
ZJU_TEMPLATE_COLUMNS = ["课程代码", "课程名称", "教师姓名", "学期", "上课时间", "上课地点", "选课时间", "选课志愿"]
ZJU_TEMPLATE_ROWS = [
    [
        "CS3100M",
        "编译原理",
        "刘老师",
        "春夏",
        "周一第3,4,5节;周三第1,2节",
        "玉泉教4-310;玉泉曹光彪西-503",
        "2025-12-19 11:59:10",
        "1.0",
    ],
    [
        "CS3221M",
        "自然语言处理导论",
        "汤老师",
        "春夏",
        "周二第3,4节{单周};周二第3,4,5节{双周};周四第1,2节",
        "玉泉教1-234;玉泉教1-234;玉泉曹光彪西-503",
        "2026-03-04 22:26:30",
        "1.0",
    ],
]
WEEKDAY_MAP = {
    "一": 1,
    "周一": 1,
    "星期一": 1,
    "二": 2,
    "周二": 2,
    "星期二": 2,
    "三": 3,
    "周三": 3,
    "星期三": 3,
    "四": 4,
    "周四": 4,
    "星期四": 4,
    "五": 5,
    "周五": 5,
    "星期五": 5,
    "六": 6,
    "周六": 6,
    "星期六": 6,
    "日": 7,
    "天": 7,
    "周日": 7,
    "周天": 7,
    "星期日": 7,
    "星期天": 7,
}


def create_course(db: Session, payload: CourseCreate, user_id: int = DEFAULT_USER_ID) -> CourseSchedule:
    validate_course_payload(payload)
    if find_existing_course(db, user_id, payload):
        raise ValueError("相同课程时段已存在，请勿重复添加。")

    course = CourseSchedule(user_id=user_id, **payload.model_dump())
    db.add(course)
    db.flush()
    db.add(build_course_event(course))
    db.commit()
    db.refresh(course)
    return course


def list_courses(db: Session, user_id: int = DEFAULT_USER_ID) -> list[dict]:
    courses = db.scalars(
        select(CourseSchedule)
        .where(CourseSchedule.user_id == user_id)
        .order_by(CourseSchedule.weekday.asc(), CourseSchedule.start_section.asc(), CourseSchedule.id.asc())
    ).all()
    return [serialize_course(course) for course in courses]


def delete_course(
    db: Session,
    course_id: int,
    user_id: int = DEFAULT_USER_ID,
    scope: str = "one",
) -> dict:
    course = db.get(CourseSchedule, course_id)
    if not course or course.user_id != user_id:
        raise ValueError("课程不存在")

    normalized_scope = normalize_delete_scope(scope)
    courses = courses_for_delete_scope(db, course, normalized_scope)
    course_ids = [item.id for item in courses]
    course_name = course.course_name

    deleted_events = 0
    for item in courses:
        deleted_events += delete_matching_course_events(db, item)
        db.delete(item)
    db.commit()
    return {
        "id": course_id,
        "scope": normalized_scope,
        "course_name": course_name,
        "deleted_courses": len(course_ids),
        "deleted_course_ids": course_ids,
        "deleted_events": deleted_events,
    }


def normalize_delete_scope(scope: str) -> str:
    aliases = {
        "one": "one",
        "single": "one",
        "once": "one",
        "day": "day",
        "same_day": "day",
        "specific_date": "day",
        "all": "all",
        "all_course": "all",
    }
    normalized = aliases.get(scope)
    if not normalized:
        raise ValueError("删除范围无效，请选择：本次、当天或全部这门课。")
    return normalized


def courses_for_delete_scope(db: Session, course: CourseSchedule, scope: str) -> list[CourseSchedule]:
    if scope == "one":
        return [course]

    stmt = select(CourseSchedule).where(
        CourseSchedule.user_id == course.user_id,
        CourseSchedule.course_name == course.course_name,
        CourseSchedule.teacher.is_(None)
        if course.teacher is None
        else CourseSchedule.teacher == course.teacher,
    )
    if scope == "day":
        stmt = stmt.where(CourseSchedule.weekday == course.weekday)

    return db.scalars(
        stmt.order_by(CourseSchedule.weekday.asc(), CourseSchedule.start_section.asc(), CourseSchedule.id.asc())
    ).all()


def import_courses_from_upload(
    db: Session,
    filename: str,
    content: bytes,
    user_id: int = DEFAULT_USER_ID,
) -> dict:
    rows = parse_course_file(filename, content)
    if not rows:
        raise ValueError("文件中没有可导入的课程行，请检查是否包含表头和至少一行课程数据。")

    courses: list[CourseSchedule] = []
    errors: list[str] = []
    seen: set[tuple] = set()

    for index, row in enumerate(rows, start=2):
        row_number = row.get("__row_number", index)
        try:
            payloads = course_payloads_from_row(row)
        except ValueError as exc:
            errors.append(f"第 {row_number} 行：{exc}")
            continue

        for payload in payloads:
            try:
                validate_course_payload(payload)
            except ValueError as exc:
                errors.append(f"第 {row_number} 行：{exc}")
                continue

            dedupe_key = (
                payload.course_name,
                payload.teacher,
                payload.weekday,
                payload.start_section,
                payload.end_section,
                payload.weeks,
                payload.location,
            )
            if dedupe_key in seen:
                errors.append(f"第 {row_number} 行：重复课程时段已跳过：{payload.course_name}")
                continue
            seen.add(dedupe_key)

            if find_existing_course(db, user_id, payload):
                errors.append(f"第 {row_number} 行：数据库中已存在该课程时段，已跳过：{payload.course_name}")
                continue

            course = CourseSchedule(user_id=user_id, **payload.model_dump())
            db.add(course)
            db.flush()
            db.add(build_course_event(course))
            courses.append(course)

    db.commit()
    for course in courses:
        db.refresh(course)

    return {
        "filename": filename,
        "imported_count": len(courses),
        "skipped_count": len(errors),
        "courses": [serialize_course(course) for course in courses],
        "errors": errors,
        "example": build_course_template_example(),
    }


def parse_course_file(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return parse_csv_courses(content)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_excel_courses(content)
    if suffix == ".xls":
        raise ValueError("暂不支持旧版 .xls 文件，请在 Excel/WPS 中另存为 .xlsx 后重新上传。")
    raise ValueError("仅支持 .csv、.xlsx 或 .xlsm 课表文件。")


def parse_csv_courses(content: bytes) -> list[dict[str, Any]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV 文件请使用 UTF-8 编码保存后重新上传。") from exc

    all_rows = list(csv.reader(io.StringIO(text)))
    if not all_rows:
        return []

    header_index, headers = find_header_row(all_rows)
    validate_headers(headers)
    rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(all_rows[header_index + 1 :], start=header_index + 2):
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
        if has_content(row):
            normalized = normalize_row(row)
            normalized["__row_number"] = row_number
            rows.append(normalized)
    return rows


def parse_excel_courses(content: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Excel 文件无法解析，请确认文件为 .xlsx/.xlsm 且未损坏。") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_index, headers = find_header_row(rows)
    validate_headers(headers)

    parsed_rows: list[dict[str, Any]] = []
    for row_number, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
        if has_content(row):
            normalized = normalize_row(row)
            normalized["__row_number"] = row_number
            parsed_rows.append(normalized)
    return parsed_rows


def find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, list[str]]:
    for index, row in enumerate(rows):
        headers = [str(value).strip() if value is not None else "" for value in row]
        if is_supported_header_set(headers):
            return index, headers
    raise ValueError(
        "未找到可识别的课表表头。支持普通模板表头“课程名,星期,节次,地点,教师,周次”，"
        "也支持教务导出表头“课程名称,教师姓名,上课时间,上课地点”。"
    )


def normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in row.items():
        canonical = canonical_header(str(key).strip())
        if canonical:
            normalized[canonical] = value.strip() if isinstance(value, str) else value
    return normalized


def canonical_header(header: str) -> str | None:
    for canonical, candidates in HEADER_MAP.items():
        if header in candidates:
            return canonical
    return None


def validate_headers(headers: list[str]) -> None:
    canonical_headers = {canonical_header(header.strip()) for header in headers if header}
    canonical_headers.discard(None)

    if is_zju_export_headers(canonical_headers):
        return
    if is_standard_headers(canonical_headers):
        return

    missing: list[str] = []
    if not REQUIRED_COLUMNS.issubset(canonical_headers):
        missing.append("课程名、星期")
    if not any(group.issubset(canonical_headers) for group in SECTION_COLUMN_GROUPS):
        missing.append("节次，或开始节次+结束节次")

    raise ValueError(
        "课表表头不完整，至少需要："
        + "；".join(missing)
        + "。示例表头：课程名,星期,节次,地点,教师,周次。"
        + "如果使用教务导出的课表，请保留“课程名称、教师姓名、上课时间、上课地点”列。"
    )


def is_supported_header_set(headers: list[str]) -> bool:
    canonical_headers = {canonical_header(header.strip()) for header in headers if header}
    canonical_headers.discard(None)
    return is_standard_headers(canonical_headers) or is_zju_export_headers(canonical_headers)


def is_standard_headers(canonical_headers: set[str | None]) -> bool:
    return REQUIRED_COLUMNS.issubset(canonical_headers) and any(
        group.issubset(canonical_headers) for group in SECTION_COLUMN_GROUPS
    )


def is_zju_export_headers(canonical_headers: set[str | None]) -> bool:
    return ZJU_EXPORT_REQUIRED_COLUMNS.issubset(canonical_headers)


def has_content(row: dict[str, Any]) -> bool:
    return any(value is not None and str(value).strip() for value in row.values())


def course_payloads_from_row(row: dict[str, Any]) -> list[CourseCreate]:
    if row.get("schedule_time"):
        return zju_export_payloads_from_row(row)

    start_section = row.get("start_section")
    end_section = row.get("end_section")
    if (not start_section or not end_section) and row.get("sections"):
        start_section, end_section = parse_sections(row["sections"])

    return [
        CourseCreate(
            course_name=required_text(row.get("course_name"), "课程名不能为空"),
            teacher=optional_text(row.get("teacher")),
            weekday=parse_weekday(row.get("weekday")),
            start_section=parse_int(start_section, "开始节次不能为空"),
            end_section=parse_int(end_section, "结束节次不能为空"),
            weeks=optional_text(row.get("weeks")),
            location=optional_text(row.get("location")),
        )
    ]


def zju_export_payloads_from_row(row: dict[str, Any]) -> list[CourseCreate]:
    course_name = required_text(row.get("course_name"), "课程名称不能为空")
    teacher = optional_text(row.get("teacher"))
    semester = optional_text(row.get("semester"))
    locations = split_locations(row.get("location"))
    slots = parse_zju_schedule_time(row.get("schedule_time"))
    if not slots:
        raise ValueError("上课时间为空或无法识别，例如：周一第3,4,5节;周三第1,2节")

    payloads: list[CourseCreate] = []
    for index, slot in enumerate(slots):
        location = locations[index] if index < len(locations) else (locations[-1] if locations else None)
        weeks = " ".join(part for part in [semester, slot["week_note"]] if part) or None
        payloads.append(
            CourseCreate(
                course_name=course_name,
                teacher=teacher,
                weekday=slot["weekday"],
                start_section=slot["start_section"],
                end_section=slot["end_section"],
                weeks=weeks,
                location=location,
            )
        )
    return payloads


def validate_course_payload(payload: CourseCreate) -> None:
    if not payload.course_name.strip():
        raise ValueError("课程名不能为空")
    if payload.weekday < 1 or payload.weekday > 7:
        raise ValueError("星期必须在 1 到 7 之间")
    if payload.start_section < 1 or payload.end_section > 13:
        raise ValueError("节次必须在 1 到 13 之间")
    if payload.start_section > payload.end_section:
        raise ValueError("开始节次不能晚于结束节次")


def build_course_event(course: CourseSchedule) -> ScheduleEvent:
    start_time, end_time = section_range_to_datetime(course.weekday, course.start_section, course.end_section)
    return ScheduleEvent(
        user_id=course.user_id,
        title=course_event_title(course.course_name),
        type="course",
        activity_id=None,
        start_time=start_time,
        end_time=end_time,
        location=course.location,
        color_type="course",
    )


def course_event_title(course_name: str) -> str:
    title = course_name.strip()
    while title.endswith(COURSE_TITLE_SUFFIX):
        title = title[: -len(COURSE_TITLE_SUFFIX)].strip()
    return title or course_name.strip()


def course_event_title_candidates(course_name: str) -> list[str]:
    normalized = course_event_title(course_name)
    candidates = [
        normalized,
        f"{normalized}{COURSE_TITLE_SUFFIX}",
        course_name.strip(),
    ]
    return list(dict.fromkeys(candidate for candidate in candidates if candidate))


def delete_matching_course_events(db: Session, course: CourseSchedule) -> int:
    start_time, end_time = section_range_to_datetime(
        course.weekday,
        course.start_section,
        course.end_section,
    )
    events = db.scalars(
        select(ScheduleEvent).where(
            ScheduleEvent.user_id == course.user_id,
            ScheduleEvent.type == "course",
            ScheduleEvent.title.in_(course_event_title_candidates(course.course_name)),
            ScheduleEvent.start_time == start_time,
            ScheduleEvent.end_time == end_time,
            ScheduleEvent.location.is_(None)
            if course.location is None
            else ScheduleEvent.location == course.location,
        )
    ).all()
    for event in events:
        db.delete(event)
    return len(events)


def section_range_to_datetime(weekday: int, start_section: int, end_section: int) -> tuple[datetime, datetime]:
    today = date.today()
    monday = today - timedelta(days=today.weekday())
    target_date = monday + timedelta(days=weekday - 1)
    start_clock = SECTION_TIMES.get(start_section, SECTION_TIMES[1])[0]
    end_clock = SECTION_TIMES.get(end_section, SECTION_TIMES[13])[1]
    return datetime.combine(target_date, start_clock), datetime.combine(target_date, end_clock)


def parse_sections(value: Any) -> tuple[int, int]:
    text = str(value).strip().replace("第", "").replace("节", "")
    for separator in ("-", "~", "—", "–", "至"):
        if separator in text:
            start, end = text.split(separator, 1)
            return parse_int(start, "开始节次不能为空"), parse_int(end, "结束节次不能为空")
    section = parse_int(text, "节次不能为空")
    return section, section


def parse_zju_schedule_time(value: Any) -> list[dict[str, int | str | None]]:
    text = required_text(value, "上课时间不能为空")
    slots: list[dict[str, int | str | None]] = []
    for segment in re.split(r"[;；]", text):
        segment = segment.strip()
        if not segment:
            continue

        match = re.match(r"周([一二三四五六日天1-7])第(.+?)节(?:\{([^{}]+)\})?$", segment)
        if not match:
            raise ValueError(f"无法识别上课时间“{segment}”，示例：周一第3,4,5节")

        weekday_text, sections_text, week_note = match.groups()
        sections = parse_section_numbers(sections_text)
        slots.append(
            {
                "weekday": parse_weekday(f"周{weekday_text}" if not weekday_text.isdigit() else weekday_text),
                "start_section": min(sections),
                "end_section": max(sections),
                "week_note": week_note,
            }
        )
    return slots


def parse_section_numbers(value: str) -> list[int]:
    normalized = (
        value.strip()
        .replace("，", ",")
        .replace("、", ",")
        .replace("~", "-")
        .replace("—", "-")
        .replace("–", "-")
        .replace("至", "-")
    )
    numbers: list[int] = []
    for part in normalized.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            start, end = part.split("-", 1)
            numbers.extend(range(parse_int(start, "节次格式错误"), parse_int(end, "节次格式错误") + 1))
        else:
            numbers.append(parse_int(part, "节次格式错误"))
    if not numbers:
        raise ValueError("节次不能为空")
    return numbers


def split_locations(value: Any) -> list[str]:
    text = optional_text(value)
    if not text:
        return []
    return [item.strip() for item in re.split(r"[;；]", text) if item.strip()]


def build_course_template_example() -> dict:
    standard_csv = "\n".join(
        [
            ",".join(COURSE_TEMPLATE_COLUMNS),
            *[",".join(row) for row in COURSE_TEMPLATE_ROWS],
        ]
    )
    zju_csv = "\n".join(
        [
            ZJU_TEMPLATE_TITLE,
            ",".join(ZJU_TEMPLATE_COLUMNS),
            *[",".join(row) for row in ZJU_TEMPLATE_ROWS],
        ]
    )
    return {
        "headers": COURSE_TEMPLATE_COLUMNS,
        "rows": COURSE_TEMPLATE_ROWS,
        "csv": standard_csv,
        "zju_headers": ZJU_TEMPLATE_COLUMNS,
        "zju_rows": ZJU_TEMPLATE_ROWS,
        "zju_title": ZJU_TEMPLATE_TITLE,
        "zju_csv": zju_csv,
        "supported_extensions": [".csv", ".xlsx", ".xlsm"],
        "notes": [
            "普通模板至少需要：课程名、星期、节次。",
            "教务导出格式可直接上传，例如课表_3230106240.xlsx。",
            "教务导出表头需包含：课程名称、教师姓名、上课时间、上课地点。",
            "多个上课时段会自动拆分，地点会按分号顺序匹配。",
        ],
    }


def parse_weekday(value: Any) -> int:
    text = required_text(value, "星期不能为空")
    if text in WEEKDAY_MAP:
        return WEEKDAY_MAP[text]
    return parse_int(text, "星期必须为 1-7 或中文星期")


def parse_int(value: Any, message: str) -> int:
    text = required_text(value, message)
    try:
        return int(float(text))
    except ValueError as exc:
        raise ValueError(message) from exc


def required_text(value: Any, message: str) -> str:
    text = "" if value is None else str(value).strip()
    if not text:
        raise ValueError(message)
    return text


def optional_text(value: Any) -> str | None:
    text = "" if value is None else str(value).strip()
    return text or None


def serialize_course(course: CourseSchedule) -> dict:
    return {
        "id": course.id,
        "course_name": course.course_name,
        "teacher": course.teacher,
        "weekday": course.weekday,
        "start_section": course.start_section,
        "end_section": course.end_section,
        "weeks": course.weeks,
        "location": course.location,
    }


def find_existing_course(
    db: Session,
    user_id: int,
    payload: CourseCreate,
) -> CourseSchedule | None:
    return db.scalar(
        select(CourseSchedule).where(
            CourseSchedule.user_id == user_id,
            CourseSchedule.course_name == payload.course_name,
            CourseSchedule.weekday == payload.weekday,
            CourseSchedule.start_section == payload.start_section,
            CourseSchedule.end_section == payload.end_section,
            CourseSchedule.teacher.is_(None)
            if payload.teacher is None
            else CourseSchedule.teacher == payload.teacher,
            CourseSchedule.location.is_(None)
            if payload.location is None
            else CourseSchedule.location == payload.location,
            CourseSchedule.weeks.is_(None)
            if payload.weeks is None
            else CourseSchedule.weeks == payload.weeks,
        )
    )
